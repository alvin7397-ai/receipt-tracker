import os, json, base64, sqlite3, io
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file, session, redirect, url_for
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'change-this-in-production')

DB_PATH   = os.environ.get('DB_PATH', os.path.join(os.path.dirname(__file__), 'expenses.db'))
API_KEY   = os.environ.get('ANTHROPIC_API_KEY', '')
APP_PASS  = os.environ.get('APP_PASSWORD', '')   # optional password protection

CATEGORIES = ['家用品', '食品', '外食費用', '貓咪食品', '貓咪用品', '訂閱費', '保險費', '收入']

CAT_META = {
    '家用品':   {'dark': '2E75B6', 'light': 'DEEAF1', 'icon': '🏠'},
    '食品':     {'dark': '548235', 'light': 'E2EFDA', 'icon': '🍱'},
    '外食費用': {'dark': '5B8DB8', 'light': 'D6E8F5', 'icon': '🍜'},
    '貓咪食品': {'dark': 'C55A11', 'light': 'FCE4D6', 'icon': '🐾'},
    '貓咪用品': {'dark': '6B3FA0', 'light': 'EDE0F5', 'icon': '🐱'},
    '訂閱費':   {'dark': '7B6B9E', 'light': 'EBE5F5', 'icon': '📱'},
    '保險費':   {'dark': '5A7A5A', 'light': 'D9EDD9', 'icon': '🛡️'},
    '收入':     {'dark': '1D6B38', 'light': 'C6EFCE', 'icon': '💰'},
}

JUNE_DATA = {
    '家用品': [
        {'date':'2026/06/22','name':'HomeSense Glass/Crystal 玻璃器皿','store':'HomeSense','amount':20.99,'note':'含GST $1.00'},
        {'date':'2026/06/22','name':'IKEA ISRANUNKEL 人體工學枕頭','store':'IKEA','amount':79.99,'note':''},
        {'date':'2026/06/22','name':'IKEA OSTBIT 盤架 x2','store':'IKEA','amount':9.98,'note':''},
        {'date':'2026/06/22','name':'IKEA GST','store':'IKEA','amount':4.50,'note':'GST 5% on $89.97'},
        {'date':'2026/06/22','name':'ecoco 多功能衛生間地刷','store':'天貓 ecoco','amount':6.26,'note':'¥29.90 CNY @ 0.2092'},
        {'date':'2026/06/22','name':'ecoco 蠔油擠壓器','store':'天貓 ecoco','amount':2.89,'note':'¥13.80 CNY @ 0.2092'},
        {'date':'2026/06/22','name':'ecoco 破壁機清潔刷','store':'天貓 ecoco','amount':3.33,'note':'¥15.90 CNY @ 0.2092'},
        {'date':'2026/06/22','name':'ecoco 杯刷三合一','store':'天貓 ecoco','amount':4.16,'note':'¥19.90 CNY @ 0.2092'},
        {'date':'2026/06/22','name':'ecoco 鞋刷洗衣刷組合','store':'天貓 ecoco','amount':5.03,'note':'¥24.04 CNY @ 0.2092'},
        {'date':'2026/06/20','name':'BLUEAIR Blue Pure 511i Max Air Purifier','store':'Amazon','amount':139.99,'note':''},
        {'date':'2026/06/17','name':'BLUEAIR Blue Pure 511i Max Air Purifier','store':'Amazon','amount':149.99,'note':''},
        {'date':'2026/06/17','name':'BLUEAIR Blue Pure 311i Max Air Purifier（換貨中）','store':'Amazon','amount':239.99,'note':'Replacement ordered'},
        {'date':'2026/06/17','name':'Levoit Evaporative Humidifier（換貨中）','store':'Amazon','amount':329.89,'note':'Replacement ordered'},
        {'date':'2026/06/17','name':'eufy Indoor Cam E30 4K Security Camera','store':'Amazon','amount':69.89,'note':''},
    ],
    '食品': [
        {'date':'2026/06/23','name':'Sunrise 甜豆漿 1.89L','store':'T&T Supermarket','amount':5.67,'note':'特價'},
        {'date':'2026/06/23','name':'T&T 保證金 + 環保費 + GST','store':'T&T Supermarket','amount':0.36,'note':'保證金$0.25 + 環保費$0.10 + GST$0.01'},
        {'date':'2026/06/23','name':'Codorniu Limited Edition Brut Cava 750ml','store':'Wine & Beyond','amount':21.98,'note':''},
        {'date':'2026/06/23','name':'Wine & Beyond 保證金 + GST','store':'Wine & Beyond','amount':1.20,'note':'保證金$0.10 + GST$1.10'},
        {'date':'2026/06/22','name':'Split Wing 雞翅','store':'Costco','amount':28.79,'note':''},
        {'date':'2026/06/22','name':'Short Ribs 牛肋排','store':'Costco','amount':40.56,'note':''},
        {'date':'2026/06/22','name':'Remedy Kombucha','store':'Costco','amount':19.99,'note':'原價$24.99 折扣-$5.00'},
        {'date':'2026/06/22','name':'Coke 32x355ml','store':'Costco','amount':17.29,'note':''},
        {'date':'2026/06/22','name':'Greek Yogurt 希臘優格','store':'Costco','amount':8.99,'note':'原價$11.49 折扣-$2.50'},
        {'date':'2026/06/22','name':'KS Tofu 豆腐','store':'Costco','amount':7.99,'note':''},
        {'date':'2026/06/22','name':'Mini Bellas 蘑菇','store':'Costco','amount':5.99,'note':''},
        {'date':'2026/06/22','name':'Chicken Lasagna 雞肉千層麵','store':'Costco','amount':19.49,'note':''},
        {'date':'2026/06/22','name':'Butter 454g 牛油','store':'Costco','amount':5.79,'note':''},
        {'date':'2026/06/22','name':'Costco 瓶罐保證金 + GST','store':'Costco','amount':6.86,'note':'保證金$5.00 + GST$1.86'},
        {'date':'2026/06/22','name':'Kronenbourg 1664 Blanc 6PK','store':'Wine & Beyond','amount':16.95,'note':'原價$20.49 折扣-$3.54'},
        {'date':'2026/06/22','name':'Chum Churum Yogurt 360ml','store':'Wine & Beyond','amount':11.99,'note':''},
        {'date':'2026/06/22','name':'Smirnoff Ice Light Peach Soda 4PK','store':'Wine & Beyond','amount':15.49,'note':''},
        {'date':'2026/06/22','name':'Wine & Beyond 保證金 + GST','store':'Wine & Beyond','amount':3.32,'note':'保證金$1.10 + GST$2.22'},
        {'date':'2026/06/22','name':'靖江猪肉脯 108g XO醬 x4袋','store':'天貓 红树林食品','amount':11.67,'note':'¥55.80 CNY @ 0.2092'},
        {'date':'2026/06/22','name':'無窮鹽焗雞翅 香辣味 x2','store':'天貓 无穷食品','amount':33.49,'note':'¥160.10 CNY @ 0.2092'},
        {'date':'2026/06/22','name':'無窮鹽焗雞翅 組合裝 x1','store':'天貓 无穷食品','amount':11.26,'note':'¥53.81 CNY @ 0.2092'},
    ],
    '外食費用': [
        {'date':'2026/06/22','name':'IKEA 餐廳 Meatballs + Cod Fritters','store':'IKEA 餐廳','amount':31.99,'note':'含GST $1.52；咖啡免費折扣-$1.50'},
    ],
    '貓咪食品': [
        {'date':'2026/06/21','name':'Inaba Churu 貓泥 3包','store':'','amount':12.58,'note':'稅前$11.98 + GST$0.60'},
        {'date':'2026/06/20','name':'Love Bugs Probiotic 40g','store':'','amount':38.84,'note':'稅前$36.99 + GST$1.85'},
        {'date':'2026/06/20','name':'Acana First Feast Dry Food','store':'','amount':31.49,'note':'稅前$29.99 + GST$1.50'},
        {'date':'2026/06/20','name':'Vital Chicken Hearts','store':'','amount':9.96,'note':'稅前$9.49 + GST$0.47'},
        {'date':'2026/06/20','name':'Vital Salmon Bites','store':'','amount':9.96,'note':'稅前$9.49 + GST$0.47'},
        {'date':'2026/06/20','name':'Kiti food 罐頭 x12','store':'','amount':26.25,'note':'稅前$25.00 + GST$1.25'},
    ],
    '貓咪用品': [
        {'date':'2026/06/21','name':'PetSmart 貓咪航空箱','store':'PetSmart','amount':52.49,'note':'稅前$49.99 + GST$2.50'},
        {'date':'2026/06/21','name':"Skout's Honor 酵素清潔液",'store':'','amount':15.11,'note':'稅前$14.39 + GST$0.72'},
        {'date':'2026/06/21','name':'FELIWAY 費洛蒙','store':'','amount':37.79,'note':'稅前$35.99 + GST$1.80'},
        {'date':'2026/06/20','name':'CHEERHOME PETS Sisal Cat Scratcher Bed','store':'Amazon','amount':43.69,'note':''},
        {'date':'2026/06/20','name':'貓鈴鐺','store':'','amount':10.49,'note':'稅前$9.99 + GST$0.50'},
        {'date':'2026/06/19','name':'Cat litter 1箱','store':'','amount':37.80,'note':'稅前$36.00 + GST$1.80'},
        {'date':'2026/06/17','name':'Pieviev Cat Litter Mat Trapper 76x61cm','store':'Amazon','amount':24.99,'note':''},
        {'date':'2026/06/17','name':'WWVVPET Pet Nail Clipper with LED Light','store':'Amazon','amount':18.95,'note':''},
        {'date':'2026/06/17','name':'aumuca Cat Brush for Shedding','store':'Amazon','amount':19.99,'note':''},
        {'date':'2026/06/17','name':'OROLEY Stainless Steel Cat Litter Box XL','store':'Amazon','amount':59.27,'note':''},
        {'date':'2026/06/17','name':'PETLIBRO Automatic Cat Feeder 2L WiFi','store':'Amazon','amount':89.99,'note':''},
        {'date':'2026/06/17','name':'PETLIBRO Cat Water Fountain Dockstream 2','store':'Amazon','amount':84.87,'note':''},
        {'date':'2026/06/22','name':'Home Sense 貓玩具','store':'HomeSense','amount':10.49,'note':'稅前$9.99 + GST$0.50'},
        {'date':'2026/06/22','name':'貓梳子 除毛梳毛刷','store':'天貓 柒哦宠物','amount':1.71,'note':'¥8.17 CNY @ 0.2092'},
        {'date':'2026/06/22','name':'Peekaboo 貓咪罐頭軟硅膠匙','store':'淘寶 Peekaboo','amount':1.35,'note':'¥6.45 CNY @ 0.2092'},
    ],
    '訂閱費': [],
    '保險費': [],
    '收入':   [],
}

# ── DB ────────────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    os.makedirs(os.path.dirname(DB_PATH) if os.path.dirname(DB_PATH) else '.', exist_ok=True)
    with get_db() as conn:
        conn.execute('''CREATE TABLE IF NOT EXISTS expenses (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            date       TEXT    NOT NULL DEFAULT '',
            name       TEXT    NOT NULL DEFAULT '',
            store      TEXT    NOT NULL DEFAULT '',
            category   TEXT    NOT NULL DEFAULT '食品',
            amount     REAL    NOT NULL DEFAULT 0,
            note       TEXT    NOT NULL DEFAULT '',
            month      TEXT    NOT NULL,
            created_at TEXT    NOT NULL DEFAULT CURRENT_TIMESTAMP
        )''')
        conn.commit()

# ── Auth ──────────────────────────────────────────────────────────────────────

@app.before_request
def check_auth():
    if not APP_PASS:
        return
    exempt = {'login', 'static'}
    if request.endpoint in exempt:
        return
    if not session.get('ok'):
        if request.is_json:
            return jsonify({'error': 'Unauthorized'}), 401
        return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        pw = (request.json or {}).get('password', '')
        if pw == APP_PASS:
            session['ok'] = True
            return jsonify({'success': True})
        return jsonify({'error': 'Wrong password'}), 401
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login') if APP_PASS else url_for('index'))

# ── Pages ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html',
                           categories=CATEGORIES,
                           cat_meta=json.dumps(CAT_META))

# ── API ───────────────────────────────────────────────────────────────────────

@app.route('/api/parse', methods=['POST'])
def parse_receipt():
    if not API_KEY:
        return jsonify({'error': 'ANTHROPIC_API_KEY not set on server'}), 500
    if 'image' not in request.files:
        return jsonify({'error': 'No image'}), 400

    f = request.files['image']
    img_bytes = f.read()
    if not img_bytes:
        return jsonify({'error': 'Empty file'}), 400

    mt = f.content_type or ''
    if not mt.startswith('image/'):
        hdr = img_bytes[:12]
        if hdr[:3] == b'\xff\xd8\xff':            mt = 'image/jpeg'
        elif hdr[:8] == b'\x89PNG\r\n\x1a\n':     mt = 'image/png'
        elif hdr[:4] == b'RIFF' and hdr[8:12] == b'WEBP': mt = 'image/webp'
        else:                                       mt = 'image/jpeg'

    img_b64 = base64.b64encode(img_bytes).decode()

    prompt = """Parse this receipt image. Return ONLY a JSON object — no markdown, no explanation.

{
  "store": "store name",
  "date": "YYYY/MM/DD",
  "items": [
    {"name": "item name", "amount": 12.99, "category": "食品", "note": ""}
  ]
}

Category rules (pick the single best match):
- 食品: groceries, food, drinks, alcohol, beverages
- 家用品: household goods, electronics, cleaning, furniture, appliances
- 外食費用: restaurant, food court, cafe (eat-in)
- 貓咪食品: cat food, cat treats, cat supplements
- 貓咪用品: cat litter, cat toys, cat accessories, pet supplies
- 訂閱費: subscription, membership, streaming
- 保險費: insurance

Additional rules:
- GST, bottle deposits, env fees → separate line items with descriptive name
- Discounted items → use final price paid; mention original price and discount in "note"
- CNY receipts → convert to CAD at 0.2092; note format: "¥XX CNY @ 0.2092"
- Date format YYYY/MM/DD; assume year 2026 if not shown
- Include ALL items including taxes and deposits"""

    try:
        client = anthropic.Anthropic(api_key=API_KEY)
        resp = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=3000,
            messages=[{'role': 'user', 'content': [
                {'type': 'image', 'source': {'type': 'base64', 'media_type': mt, 'data': img_b64}},
                {'type': 'text', 'text': prompt}
            ]}]
        )
        text = resp.content[0].text.strip()
        # Strip markdown code fences
        if '```' in text:
            for part in text.split('```'):
                part = part.strip()
                if part.startswith('json'):
                    part = part[4:].strip()
                try:
                    return jsonify(json.loads(part))
                except Exception:
                    pass
        return jsonify(json.loads(text))
    except json.JSONDecodeError:
        return jsonify({'error': 'AI returned invalid JSON', 'raw': text}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/entries', methods=['GET'])
def list_entries():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    with get_db() as conn:
        rows = conn.execute(
            'SELECT * FROM expenses WHERE month=? ORDER BY date DESC, id DESC', (month,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/entries', methods=['POST'])
def create_entries():
    data   = request.json or {}
    entries = data.get('entries', [])
    month  = data.get('month', datetime.now().strftime('%Y-%m'))
    ids = []
    with get_db() as conn:
        for e in entries:
            cur = conn.execute(
                'INSERT INTO expenses (date,name,store,category,amount,note,month) VALUES (?,?,?,?,?,?,?)',
                (e.get('date',''), e.get('name',''), e.get('store',''),
                 e.get('category','食品'), float(e.get('amount') or 0),
                 e.get('note',''), month)
            )
            ids.append(cur.lastrowid)
        conn.commit()
    return jsonify({'success': True, 'count': len(ids), 'ids': ids})


@app.route('/api/entries/<int:eid>', methods=['PUT'])
def update_entry(eid):
    e = request.json or {}
    with get_db() as conn:
        conn.execute(
            'UPDATE expenses SET date=?,name=?,store=?,category=?,amount=?,note=? WHERE id=?',
            (e.get('date',''), e.get('name',''), e.get('store',''),
             e.get('category','食品'), float(e.get('amount') or 0), e.get('note',''), eid)
        )
        conn.commit()
    return jsonify({'success': True})


@app.route('/api/entries/<int:eid>', methods=['DELETE'])
def delete_entry(eid):
    with get_db() as conn:
        conn.execute('DELETE FROM expenses WHERE id=?', (eid,))
        conn.commit()
    return jsonify({'success': True})


@app.route('/api/months')
def list_months():
    with get_db() as conn:
        rows = conn.execute(
            'SELECT DISTINCT month FROM expenses ORDER BY month DESC'
        ).fetchall()
    months = [r['month'] for r in rows]
    if not months:
        months = [datetime.now().strftime('%Y-%m')]
    return jsonify(months)


@app.route('/api/import-june', methods=['POST'])
def import_june():
    month = '2026-06'
    with get_db() as conn:
        existing = conn.execute(
            'SELECT COUNT(*) FROM expenses WHERE month=?', (month,)
        ).fetchone()[0]
        if existing > 0:
            return jsonify({'error': f'Already have {existing} entries for June 2026'}), 409
        count = 0
        for cat, entries in JUNE_DATA.items():
            for e in entries:
                conn.execute(
                    'INSERT INTO expenses (date,name,store,category,amount,note,month) VALUES (?,?,?,?,?,?,?)',
                    (e['date'], e['name'], e.get('store',''), cat,
                     float(e['amount']), e.get('note',''), month)
                )
                count += 1
        conn.commit()
    return jsonify({'success': True, 'count': count})


# ── Excel export ──────────────────────────────────────────────────────────────

@app.route('/api/export')
def export_excel():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    with get_db() as conn:
        rows = conn.execute(
            'SELECT * FROM expenses WHERE month=? ORDER BY category, date, id', (month,)
        ).fetchall()

    data = {cat: [] for cat in CATEGORIES}
    for row in rows:
        cat = row['category']
        if cat in data:
            data[cat].append(dict(row))

    NAVY = '1F3864'
    C = {
        'navy': NAVY, 'blue': '2E75B6', 'blue_lt': 'DEEAF1',
        'green': '548235', 'green_lt': 'E2EFDA',
        'orange': 'C55A11', 'orange_lt': 'FCE4D6',
        'purple': '6B3FA0', 'purple_lt': 'EDE0F5',
        'white': 'FFFFFF',
    }

    def fill(c):  return PatternFill('solid', start_color=c)
    def bdr():
        s = Side(style='thin', color='D0D0D0')
        return Border(left=s, right=s, top=s, bottom=s)
    def al(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def fn(bold=False, size=10, color='000000'):
        return Font(name='Arial', bold=bold, size=size, color=color)

    EXPENSE_CATS = [
        ('家用品',   C['blue'],   C['blue_lt'],   '🏠'),
        ('食品',     C['green'],  C['green_lt'],  '🍱'),
        ('外食費用', '5B8DB8',    'D6E8F5',       '🍜'),
        ('貓咪食品', C['orange'], C['orange_lt'], '🐾'),
        ('貓咪用品', C['purple'], C['purple_lt'], '🐱'),
        ('訂閱費',   '7B6B9E',   'EBE5F5',        '📱'),
        ('保險費',   '5A7A5A',   'D9EDD9',        '🛡️'),
    ]
    INC = ('收入', '1D6B38', 'C6EFCE', '💰')

    wb = openpyxl.Workbook()
    wb.calculation.forceFullCalc = True
    year, mon = month.split('-')
    ws = wb.active
    ws.title = f'{year}年{mon}月'

    ws.merge_cells('A1:F1')
    ws['A1'] = f'月度開支表  —  {year} 年 {mon} 月'
    ws['A1'].font = fn(bold=True, size=16, color=NAVY)
    ws['A1'].fill = fill('EBF0FA')
    ws['A1'].alignment = al('center')
    ws.row_dimensions[1].height = 38

    for ci, (h, w) in enumerate(zip(
        ['日期','品項名稱','商店 / 來源','分類','金額 (CAD)','備註'],
        [11, 36, 18, 10, 14, 20]
    ), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = fn(bold=True, color=C['white'])
        c.fill = fill(NAVY); c.alignment = al('center'); c.border = bdr()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    row_cur = 3
    sub_refs = []

    for cat_label, dark, light, icon in EXPENSE_CATS:
        entries = data.get(cat_label, [])
        n = max(len(entries), 5)

        ws.merge_cells(f'A{row_cur}:F{row_cur}')
        c = ws.cell(row=row_cur, column=1, value=f'{icon}  {cat_label}')
        c.font = fn(bold=True, size=12, color=C['white'])
        c.fill = fill(dark); c.alignment = al('left'); c.border = bdr()
        ws.row_dimensions[row_cur].height = 22
        row_cur += 1
        first = row_cur

        for i in range(n):
            r = row_cur + i
            bg = light if i % 2 == 0 else C['white']
            e = entries[i] if i < len(entries) else {}

            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = fill(bg)
                ws.cell(row=r, column=col).border = bdr()
                ws.cell(row=r, column=col).font = fn()

            ws.cell(row=r, column=1).value = e.get('date','')
            ws.cell(row=r, column=1).alignment = al('center')
            ws.cell(row=r, column=2).value = e.get('name','')
            ws.cell(row=r, column=2).alignment = al('left')
            ws.cell(row=r, column=3).value = e.get('store','')
            ws.cell(row=r, column=3).alignment = al('left')
            ws.cell(row=r, column=4).value = cat_label
            ws.cell(row=r, column=4).font = fn(color='888888')
            ws.cell(row=r, column=4).alignment = al('center')
            amt = ws.cell(row=r, column=5)
            amt.value = e.get('amount') or None
            amt.alignment = al('right'); amt.number_format = '#,##0.00'
            ws.cell(row=r, column=6).value = e.get('note','')
            ws.cell(row=r, column=6).alignment = al('left')
            ws.row_dimensions[r].height = 16

        last = row_cur + n - 1
        row_cur += n

        ws.merge_cells(f'A{row_cur}:D{row_cur}')
        c = ws.cell(row=row_cur, column=1, value=f'{cat_label} 小計')
        c.font = fn(bold=True, color=C['white']); c.fill = fill(dark)
        c.alignment = al('right'); c.border = bdr()
        sub = ws.cell(row=row_cur, column=5, value=f'=SUM(E{first}:E{last})')
        sub.font = fn(bold=True, color=C['white']); sub.fill = fill(dark)
        sub.alignment = al('right'); sub.number_format = '#,##0.00'; sub.border = bdr()
        ws.cell(row=row_cur, column=6).fill = fill(dark)
        ws.cell(row=row_cur, column=6).border = bdr()
        ws.row_dimensions[row_cur].height = 20
        sub_refs.append(f'E{row_cur}')
        row_cur += 2

    # Expense total
    exp_row = row_cur
    ws.merge_cells(f'A{exp_row}:D{exp_row}')
    c = ws.cell(row=exp_row, column=1, value='支出合計')
    c.font = fn(bold=True, size=12, color=C['white']); c.fill = fill(NAVY)
    c.alignment = al('right'); c.border = bdr()
    et = ws.cell(row=exp_row, column=5, value='='+'+'.join(sub_refs))
    et.font = fn(bold=True, size=12, color=C['white']); et.fill = fill(NAVY)
    et.alignment = al('right'); et.number_format = '#,##0.00'; et.border = bdr()
    ws.cell(row=exp_row, column=6).fill = fill(NAVY)
    ws.cell(row=exp_row, column=6).border = bdr()
    ws.row_dimensions[exp_row].height = 24
    row_cur += 2

    # Income
    inc_label, inc_dark, inc_light, inc_icon = INC
    inc_entries = data.get('收入', [])
    inc_n = max(len(inc_entries), 5)

    ws.merge_cells(f'A{row_cur}:F{row_cur}')
    c = ws.cell(row=row_cur, column=1, value=f'{inc_icon}  {inc_label}')
    c.font = fn(bold=True, size=12, color=C['white']); c.fill = fill(inc_dark)
    c.alignment = al('left'); c.border = bdr()
    ws.row_dimensions[row_cur].height = 22
    row_cur += 1
    inc_first = row_cur

    for i in range(inc_n):
        r = row_cur + i
        bg = inc_light if i % 2 == 0 else C['white']
        e = inc_entries[i] if i < len(inc_entries) else {}
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = fill(bg)
            ws.cell(row=r, column=col).border = bdr()
            ws.cell(row=r, column=col).font = fn()
        ws.cell(row=r, column=1).value = e.get('date','')
        ws.cell(row=r, column=1).alignment = al('center')
        ws.cell(row=r, column=2).value = e.get('name','')
        ws.cell(row=r, column=3).value = e.get('store','')
        ws.cell(row=r, column=4).value = '收入'
        ws.cell(row=r, column=4).font = fn(color='888888')
        ws.cell(row=r, column=4).alignment = al('center')
        amt = ws.cell(row=r, column=5)
        amt.value = e.get('amount') or None
        amt.alignment = al('right'); amt.number_format = '#,##0.00'
        ws.cell(row=r, column=6).value = e.get('note','')
        ws.row_dimensions[r].height = 16

    inc_last = row_cur + inc_n - 1
    row_cur += inc_n

    ws.merge_cells(f'A{row_cur}:D{row_cur}')
    c = ws.cell(row=row_cur, column=1, value='收入合計')
    c.font = fn(bold=True, size=12, color=C['white']); c.fill = fill(inc_dark)
    c.alignment = al('right'); c.border = bdr()
    inc_sub = ws.cell(row=row_cur, column=5, value=f'=SUM(E{inc_first}:E{inc_last})')
    inc_sub.font = fn(bold=True, size=12, color=C['white']); inc_sub.fill = fill(inc_dark)
    inc_sub.alignment = al('right'); inc_sub.number_format = '#,##0.00'; inc_sub.border = bdr()
    ws.cell(row=row_cur, column=6).fill = fill(inc_dark)
    ws.cell(row=row_cur, column=6).border = bdr()
    ws.row_dimensions[row_cur].height = 24
    inc_ref = f'E{row_cur}'
    row_cur += 2

    # Net
    ws.merge_cells(f'A{row_cur}:D{row_cur}')
    c = ws.cell(row=row_cur, column=1, value='月度淨收支')
    c.font = fn(bold=True, size=13, color=C['white']); c.fill = fill(NAVY)
    c.alignment = al('right'); c.border = bdr()
    net = ws.cell(row=row_cur, column=5, value=f'={inc_ref}-E{exp_row}')
    net.font = fn(bold=True, size=13, color=C['white']); net.fill = fill(NAVY)
    net.alignment = al('right'); net.number_format = '#,##0.00'; net.border = bdr()
    ws.cell(row=row_cur, column=6).fill = fill(NAVY)
    ws.cell(row=row_cur, column=6).border = bdr()
    ws.row_dimensions[row_cur].height = 28
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'開支表_{year}年{mon}月.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5001))
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('DEBUG','').lower() == 'true')
